
from django.urls import reverse_lazy
from .models import MenuType
from .forms import MenuTypeForm
from .forms import OrganizationForm
from user.permission import IsAdminMixin, SearchMixin
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    UpdateView,
    View,
    TemplateView,

)

from alice_menu.utils import DeleteMixin

from django.shortcuts import render


from django.contrib.auth.mixins import LoginRequiredMixin

class IndexView(LoginRequiredMixin, TemplateView):
    login_url = '/login/'
    template_name = "index.html"


class MenuTypeMixin(IsAdminMixin):
    model = MenuType
    form_class = MenuTypeForm
    paginate_by = 50    
    queryset = MenuType.objects.all()
    success_url = reverse_lazy("menu_type_list")
    search_lookup_fields = [
        "title",
        "description",
    ]


class MenuTypeList(MenuTypeMixin, ListView):
    template_name = "menutype/menutype-list.html"
    queryset = MenuType.objects.all()


class MenuTypeDetail(MenuTypeMixin, DetailView):
    template_name = "menutype/menutype-list.html"


class MenuTypeCreate(MenuTypeMixin, CreateView):
    template_name = "create.html"


class MenuTypeUpdate(MenuTypeMixin, UpdateView):
    template_name = "update.html"


class MenuTypeDelete(MenuTypeMixin, DeleteMixin, View):
    pass


from django.views.generic import ListView
from .models import Menu, MenuType
from itertools import chain
from django.db.models import Q



    
from .forms import MenuForm


class MenuMixin(IsAdminMixin):
    model = Menu
    form_class = MenuForm
    paginate_by = 50
    queryset = Menu.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("menu_list")
    search_lookup_fields = [
        "item_name",
        "menutype__title",
    ]

from .models import FlagMenu
class MenuList(MenuMixin, ListView):
    template_name = "menu/menu-list.html"
    queryset = Menu.objects.filter(status=True, is_deleted=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['flag_menu_active'] = FlagMenu.objects.first().use_same_menu_for_multiple_outlet
        context['auto_order_active'] = FlagMenu.objects.first().autoaccept_order
        return context

class MenuDetail(MenuMixin, DetailView):
    template_name = "menu/menu-detail.html"


class MenuCreate(MenuMixin, CreateView):
    template_name = "menu/menu-create.html"


class MenuUpdate(MenuMixin, UpdateView):
    template_name = "update.html"
    

class MenuDelete(MenuMixin, DeleteMixin, View):
    pass

from django.conf import settings
from django.db import IntegrityError
from django.views import View
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.contrib import messages
from openpyxl import load_workbook
from .models import Menu, MenuType
from urllib.parse import urlparse, urlunparse
import requests
from django.core.files.base import ContentFile

class MenuUpload(View):

    def post(self, request):
        file = request.FILES.get('file', None)
        if not file:
            messages.error(request, 'Please Provide the correct file ')
            return redirect(reverse_lazy("menu_create"))
        
        file_ext = file.name.split('.')[-1]
        if file_ext not in ['xlsx', 'xls']:
            messages.error(request, 'Format must be in xlsx or xls ')
            return redirect(reverse_lazy("menu_create"))

        wb = load_workbook(file)
        excel_data = list()
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(cell.value)
                row_data.insert(0, sheet.title)
                excel_data.append(row_data)
        
        for data in excel_data:
            if not all(data[0:9]):
                continue
            if data[1].lower() == 'group':
                continue
            try:
                product = Menu.objects.get(item_name__iexact=data[2].strip())
                product.group = data[1].strip()
                product.price = data[3]
                product.save()
            except Menu.DoesNotExist:
                product = Menu()
                product.group = data[1].strip()
                product.item_name=data[2].strip()
                product.price = data[3]

                product.save()
        return redirect(reverse_lazy("menu_list"))


class MenuTypeProductList(ListView):
    template_name = "menupreset/menu_preset_list.html"
    queryset = Menu.objects.filter(status=True, is_deleted=False)

    def get(self, request, *args, **kwargs):
        menu_preset_id = kwargs.get('id')
        menutype = MenuType.objects.get(id=menu_preset_id)
        menus_in_menu_preset = Menu.objects.filter(status=True, is_deleted=False, menutype=menutype)

        context = {'object_list':menus_in_menu_preset, 'menutype': menutype.title, 'menutype_id':menutype.id}

        return render(request, self.template_name, context)

from menu.models import MenuType
from django.views.generic.edit import CreateView
import json
from django.contrib import messages
from django.urls import reverse
class MenuTypeProductCreate(MenuTypeMixin, CreateView):
    queryset = MenuType.objects.all()
    template_name = "menupreset/menu_preset_create.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # token = env('API_TOKEN')
        context['products'] = Menu.objects.filter(is_deleted=False, status=True)  # Replace with actual products
        # context['token'] = token
        return context


    def post(self, request, *args, **kwargs):
        menutype_id = kwargs.get('id')
        print("The menutype id is ", menutype_id)
        selected_products_json = request.POST.get('selectedProducts')
        selected_products = json.loads(selected_products_json)
        # print("I am here")
        print(selected_products_json)
        print(selected_products)

        for product in selected_products:
            selected_product_id = product['selectedProductId']
            selected_product_name = product['selectedProductName']

            # if FeaturedProducts.objects.filter(product_id=selected_product_id, is_featured=True, is_deleted=False).exists():
            #     messages.warning(request, f"Product '{selected_product_name}' is already featured.")
            # else:
            product = Menu.objects.get(id=selected_product_id)
            menutype = MenuType.objects.get(pk=menutype_id)
            product.menutype = menutype
            product.save()
        
        return redirect(reverse('menu_preset_product_list', kwargs={'id': menutype_id}))


class MenuTypeProductDelete(MenuTypeMixin, View):
    def get(self, request, *args, **kwargs):
        product_id = kwargs.get('id')
        menutype_id = kwargs.get('menutype_id')
        print(product_id)
        menu = Menu.objects.get(pk=product_id)
        menu.menutype = None
        menu.save()

        return redirect(reverse('menu_preset_product_list', kwargs={'id': menutype_id}))

from .models import Organization 
class OrganizationMixin(IsAdminMixin):
    model = Organization
    form_class = OrganizationForm
    paginate_by = 50
    queryset = Organization.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("menu_type_list")


class OrganizationDetail(OrganizationMixin, DetailView):
    template_name = "organization/organization_detail.html"

    def get_object(self):
        return Organization.objects.last()

    # def get_context_data(self, **kwargs):
    #     context = super().get_context_data(**kwargs)
    #     users = User.objects.all()
    #     branches = Branch.objects.filter(
    #         is_deleted=False, organization=self.get_object().id
    #     )
    #     customers = Customer.objects.filter(is_deleted=False)
    #     context["branches"] = branches
    #     context["customers"] = customers
    #     context["users"] = users

    #     return context


class OrganizationCreate(OrganizationMixin, CreateView):
    template_name = "create.html"


class OrganizationUpdate(OrganizationMixin, UpdateView):
    template_name = "update.html"

    def get_object(self):
        return Organization.objects.last()
    

class OrganizationDelete(OrganizationMixin, DeleteMixin, View):
    pass